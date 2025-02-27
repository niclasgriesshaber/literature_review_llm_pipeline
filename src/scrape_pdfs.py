#!/usr/bin/env python3
"""
Script 1: Scrape (download) PDFs from Excel links, saving them into data/pdfs/.
No logs, no advanced error handling, minimal code.
"""

import os
import sys
import requests
import openpyxl
from pathlib import Path
import re

def main():
    root_dir = Path(__file__).resolve().parents[1]
    excel_path = root_dir / "data" / "deepresearch_review.xlsx"
    pdfs_dir = root_dir / "data" / "pdfs"

    # Create pdfs folder if not existing
    pdfs_dir.mkdir(exist_ok=True, parents=True)

    # Load Excel
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active  # or wb['SheetName'] if you have a named sheet

    # Debugging: Print column headers to verify structure
    first_row = list(next(sheet.iter_rows(values_only=True)))
    print(f"First row: {first_row}")

    # Determine column indices based on headers
    try:
        title_idx = first_row.index('Title')
        year_idx = first_row.index('Year')
        link_idx = first_row.index('Link')
        print(f"Found columns - Title: {title_idx}, Year: {year_idx}, Link: {link_idx}")
    except ValueError as e:
        print(f"Error finding column headers: {e}")
        print("Make sure your Excel file has 'Title', 'Year', and 'Link' columns")
        return

    download_count = 0
    skip_count = 0

    # Iterate rows, skipping header row
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row_idx == 1:
            continue  # skip header

        # Make sure we have enough columns
        if len(row) <= max(title_idx, year_idx, link_idx):
            print(f"Row {row_idx} doesn't have enough columns: {row}")
            continue

        title = row[title_idx]
        year = row[year_idx]
        link = row[link_idx]

        # If link is missing or not a string, skip
        if not link or not isinstance(link, str):
            print(f"Row {row_idx}: Missing or invalid link")
            skip_count += 1
            continue

        # Transform arxiv links to PDF links
        if "arxiv.org/abs/" in link:
            # Extract the ID using regex
            arxiv_id_match = re.search(r'arxiv\.org/abs/([^/]+)', link)
            if arxiv_id_match:
                arxiv_id = arxiv_id_match.group(1)
                link = f"https://arxiv.org/pdf/{arxiv_id}.pdf"
                print(f"Transformed to PDF link: {link}")

        # Derive a filename from link and title
        if link.split("/")[-1].strip() and '.' in link.split("/")[-1]:
            filename = link.split("/")[-1]
        else:
            # Use title as fallback if URL doesn't have a good filename
            safe_title = re.sub(r'[^\w\-_]', '_', title if title else f"paper_{row_idx}")
            filename = f"{safe_title}_{year}.pdf"
            
        if not filename.endswith(".pdf"):
            filename += ".pdf"

        pdf_path = pdfs_dir / filename

        # Skip download if already exists
        if pdf_path.is_file():
            print(f"Already have PDF: {filename}, skipping.")
            skip_count += 1
            continue

        # Download the PDF
        try:
            print(f"Downloading {link} -> {filename}")
            resp = requests.get(link, timeout=60)
            resp.raise_for_status()

            with open(pdf_path, "wb") as f:
                f.write(resp.content)

            print(f"Saved PDF: {pdf_path}")
            download_count += 1
        except Exception as e:
            print(f"Error downloading from {link}: {e}")

    print(f"Done. Downloaded {download_count} PDFs, skipped {skip_count}.")

if __name__ == "__main__":
    main()